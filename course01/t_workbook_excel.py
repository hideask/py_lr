from openpyxl import Workbook

# 假设你有一个字典
workbook_data = {'序号': [1, 2, 3, 4, 5],
                 '姓名': ['王润仙', '陈英', '刘芬明', '林禄芳', '周素蓉'],
                 '身份证号': ['432565194710111107', '244305198610301955', '244305198610301955', '814205195405243204', '814205195405243204'],
                 '离退休日期': ['20071101', '20061009', '20061009', '20091130', '20091130'],
                 '离退休标志': ['是', '是', '是', '是', '是'], '生存状态': ['正常', '正常', '正常', '正常', '正常'],
                 '待遇享受状态': ['正常发放', '正常发放', '正常发放', '正常发放', '正常发放'],
                 '当前待遇金额': ['2994.45', '3191.54', '1362.82', '2391.27', '1216.88']}

r = ['3012679785','3012794513','3012708602','3015323810']
r.__len__()
# 创建一个新的工作簿
wb = Workbook()

# 选择活动工作表
ws = wb.active

# 写入标题行
for col_num, key in enumerate(workbook_data.keys(), 1):
    ws.cell(row=1, column=col_num, value=key)


# 写入数据行
for row_num, row_data in enumerate(zip(*(workbook_data.values())), 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

    # 保存工作簿
wb.save("F:/data/data.xlsx")