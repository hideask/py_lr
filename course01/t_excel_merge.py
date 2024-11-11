import os
import time
from pathlib import Path
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Border, Side,Alignment
import shutil
import copy
from openpyxl.utils import get_column_letter,column_index_from_string,range_boundaries
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor,TwoCellAnchor
from openpyxl.drawing.xdr import  XDRPositiveSize2D
from openpyxl.worksheet.datavalidation import DataValidation


if __name__ == '__main__' :
    excelDirPath = 'F:\data\RiskList_02'
    excelDirWorkPath = 'F:\data\RiskList_02\InfoOut'
    excelDirErrorPath = ''

    #part 1
    try:
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 开始初始化读取文件路径")
        excelDirWorkPath=os.path.join(excelDirPath, "work")
        folder_path_obj = Path(excelDirWorkPath)

        # 如果 work文件夹已经存在 则需要删除文件夹及其下属文件避免对后续处理产生影响
        if folder_path_obj.exists():
            shutil.rmtree(folder_path_obj)
        # 创建文件夹（如果路径不存在，则报错）
        os.makedirs(excelDirWorkPath)

        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 开始初始化错误文件路径")
        excelDirErrorPath = os.path.join(excelDirPath, "error")
        folder_path_obj = Path(excelDirErrorPath)

        # 如果 error文件夹已经存在 则需要删除文件夹及其下属文件避免对后续处理产生影响
        if folder_path_obj.exists():
            shutil.rmtree(folder_path_obj)
        # 创建文件夹（如果路径不存在，则报错）
        os.makedirs(excelDirErrorPath)
    except FileNotFoundError:
        print("Error: The specified path does not exist")


    #part 2
    print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 开始逐一向系统级隐患清单添加项目名称列并预处理")
    # 列出文件夹下的所有文件和文件夹
    filenames = os.listdir(excelDirPath)

    #处理功能缺陷是否符合字段
    def func_1(fun_defect, fun_defect_map, index, name):
        tempt = fun_defect[index]
        if tempt is not None and tempt not in fun_defect_map[name][0]:
            fun_defect_map[name][0].append(tempt)
        if tempt in ['不符合','部分符合']:
            return 1
        else:
            return 0

    #处理功能缺陷字符
    def func_2(fun_defect, fun_defect_map, index, name, index2, accord_index, match_array):
        # print('序号：' + name + '；index1:' + str(index) + '；accord_index:' + str(accord_index))
        accord = fun_defect[accord_index]
        # if accord == '不符合' or accord == '部分符合':
        if accord in match_array:
            fun_defect_map[name][index2] +=  (fun_defect[3] + '：' + ('无' if fun_defect[index] is None else str(fun_defect[index])) + '；\n')

    # 处理用户同意改进措施
    def func_3(fun_defect, fun_defect_map, index, name):
        tempt = fun_defect[index]
        if tempt is not None and tempt not in fun_defect_map[name][5]:
            fun_defect_map[name][5].append(tempt)

        # 处理功能缺陷是否符合字段

    def func_4(fun_defect, fun_defect_map, index, name, index2):
        tempt = fun_defect[index]
        if tempt == '无需整改':
            tempt = '无须整改'
        if tempt is not None and tempt not in fun_defect_map[name][index2]:
            fun_defect_map[name][index2].append(tempt)
        if tempt in ['整改中','部分整改']:
            return 1
        else:
            return 0

    def func_5(fun_defect, index):
        # print(fun_defect)
        # print(index)
        return 0 if fun_defect[index] is None else int(fun_defect[index])

    def all_values_in_set(lst, valid_set):
        return all(value in valid_set for value in lst)

    def move_data_validation_to_next_column(worksheet):
        # 获取当前工作表的所有数据验证对象
        data_validations = worksheet.data_validations.dataValidation
        #
        # # 遍历每一列
        # for col in worksheet.columns:
        #     # 遍历所有的数据验证对象
        #     for dv in list(data_validations):
        #         # 检查数据验证的范围是否包含当前列
        #         if any(col[0].coordinate in cell_range for cell_range in dv.sqref):
        #             # 如果是，则获取数据验证的设置
        #             validation_type = dv.type
        #             showDropDown = dv.showDropDown
        #             formula1 = dv.formula1
        #             prompt = dv.prompt
        #             promptTitle = dv.promptTitle
        #             showErrorMessage = dv.showErrorMessage
        #             showInputMessage = dv.showInputMessage
        #             # 获取数据验证的范围
        #             sqref = dv.sqref
        #             # 分割范围字符串（假设它是用空格分隔的多个范围）
        #             sqref_list = sqref[0].coord.split(':')
        #
        #             # 修改数据验证的范围以指向下一列
        #             next_col_letters = [
        #                 get_column_letter(column_index_from_string(col_letter) + 1) for
        #                 col_letter in sqref_list]
        #             next_sqref = ' '.join(next_col_letters)
        #
        #             # 从工作表中移除当前的数据验证对象
        #             worksheet.remove_data_validation(dv)
        #
        #             # 创建一个新的数据验证对象并应用到下一列
        #             new_dv = DataValidation(type=validation_type, showDropDown=showDropDown, formula1=formula1,
        #                                     prompt=prompt, promptTitle=promptTitle, showErrorMessage=showErrorMessage,
        #                                     showInputMessage=showInputMessage, sqref=next_sqref)
        #             worksheet.add_data_validation(new_dv)
        # 遍历每一个数据验证对象
        for dv in data_validations:
            # 获取数据验证的范围
            sqref = dv.sqref

            # 遍历每一个范围并更新其列引用
            new_ranges = []
            for rng in sqref:
                # 假设范围是一个简单的单元格范围，如 'A1:A10'
                col_start, row_start, col_end, row_end =  rng.min_col, rng.min_row, rng.max_col, rng.max_row

                # 计算新列的字母表示
                new_col_start = get_column_letter(col_start + 1)
                new_col_end = get_column_letter(col_end + 1) if col_start != col_end else new_col_start

                # 创建新的范围字符串
                new_range = f'{new_col_start}{row_start}:{new_col_end}{row_end}'

                # 添加到新范围列表中
                new_ranges.append(new_range)

            # 构建新的 sqref 字符串
            new_sqref = ' '.join(new_ranges)

            # 创建新的数据验证对象并应用
            new_dv = DataValidation(type=dv.type, showDropDown=dv.showDropDown, formula1=dv.formula1,
                                    showErrorMessage=dv.showErrorMessage, prompt=dv.prompt, promptTitle=dv.promptTitle,
                                    showInputMessage=dv.showInputMessage, sqref=new_sqref)

            # 添加新的数据验证到工作表（这将自动移除旧的）
            worksheet.add_data_validation(new_dv)

            # 从工作表中移除当前的数据验证对象
            worksheet.remove_data_validation(dv)

    # 遍历文件夹下的所有文件
    for filename in filenames:
        if not filename.endswith(".xlsx"):
            continue

        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 开始预处理文件 " + filename)
        file_path = os.path.join(excelDirPath, filename)
        save_path = os.path.join(excelDirWorkPath, filename)
        save_error_path = os.path.join(excelDirErrorPath, filename)
        # try:
        projectName = filename.split("_")[0]+"_"+filename.split("_")[1]
        # 读取excel中的第一个sheet
        wb = load_workbook(file_path)
        sheet = wb[wb.sheetnames[0]]
        maxRow = sheet.max_row

        # 记录与原有的列宽 以便于后续使用
        colWidths = []
        for col_dimension in sheet.column_dimensions:
            if column_index_from_string(col_dimension) > 2:
                colWidths.append(sheet.column_dimensions[col_dimension].width)
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 插入项目列")
        # 插入列后可能会影响已有的合并单元格操作因此需要记录以方便后续还原 以避免影响后续操作
        # 记录所有合并单元格
        mergeCellArray = []
        for coordinate in sheet.merged_cells:
            mergeCellArray.append(coordinate.coord)
        # 取消合并单元格   取消合并和读取分开 否则会缺少内容
        for coordinate in mergeCellArray:
            sheet.unmerge_cells(coordinate)
        #获取每列下拉筛选
        # validations = sheet.data_validations.dataValidation

        #  插入空白列
        maxCol = sheet.max_column
        _insertColumn = 2

        # 把功能性缺陷合并到系统缺陷
        sheet2 = wb[wb.sheetnames[1]]
        maxRow2 = sheet2.max_row
        maxCol2 = sheet2.max_column

        fun_defect_map = {'name': '',
                          '22（用扩）':[[],0,'','','',[],'',[],[],'','',0,''],'23（用扩）':[[],0,'','','',[],'',[],[],'','',0,''],
                          '24':[[],0,'','','',[],'',[],[],'','',0,''],'25':[[],0,'','','',[],'',[],[],'','',0,''],
                          '26':[[],0,'','','',[],'',[],[],'','',0,''],'30（用扩）':[[],0,'','','',[],'',[],[],'','',0,''],
                          '33（银扩）':[[],0,'','','',[],'',[],[],'','',0,''],'34（银扩）':[[],0,'','','',[],'',[],[],'','',0,''],
                          '35':[[],0,'','','',[],'',[],[],'','',0,''],'36':[[],0,'','','',[],'',[],[],'','',0,''],
                          '37':[[],0,'','','',[],'',[],[],'','',0,''],'38':[[],0,'','','',[],'',[],[],'','',0,''],
                          '51':[[],0,'','','',[],'',[],[],'','',0,'']}

        fun_defect_list = []
        for _row2 in range(5, maxRow2, 1):
            # for _col2 in range(maxCol2, 0, -1):
            row_datas = sheet2[_row2]
            r_data_list = []
            for r_data in row_datas:
                r_data_list.append(r_data.value)
            if r_data_list[3] is None or r_data_list[0] == 0:
                continue
            fun_defect_list.append(r_data_list)
                # source_cell2 = sheet2.cell(row=_row2, column=_col2)

        for fun_defect in fun_defect_list:
            #拼接功能名称
            fun_defect_map['name'] += (fun_defect[3] + '；' if fun_defect[3] is not None else '')
            index1 = 4 #是否符合
            index2 = 6 #证明材料
            index3 = 7 #改进措施
            index4 = 8 #用户同意改进措施
            index5 = 9 #封堵结果
            index6 = 12 #审核结果
            index7 = 10 #封堵材料
            index8 = 11 #整改风险点数
            no_num = 0
            d_num = 0
            #处理'22（用扩）
            no_num = func_1(fun_defect,fun_defect_map,index1,'22（用扩）')
            fun_defect_map['22（用扩）'][1] += no_num
            func_2(fun_defect, fun_defect_map, index2, '22（用扩）', 2, index1, ['不符合','部分符合'])
            func_2(fun_defect, fun_defect_map, index3, '22（用扩）', 3, index1, ['不符合','部分符合'])
            func_3(fun_defect, fun_defect_map, index4, '22（用扩）')
            d_num = func_4(fun_defect, fun_defect_map, index5, '22（用扩）', 7)
            d_num = func_5(fun_defect,index8)
            fun_defect_map['22（用扩）'][11] += d_num
            func_4(fun_defect, fun_defect_map, index6, '22（用扩）', 8)
            func_2(fun_defect, fun_defect_map, index7, '22（用扩）', 12, index5, ['整改中','部分整改'])


            #处理'23（用扩）
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '23（用扩）')
            fun_defect_map['23（用扩）'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '23（用扩）', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '23（用扩）', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '23（用扩）')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '23（用扩）', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['23（用扩）'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '23（用扩）', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '23（用扩）', 12, index5, ['整改中','部分整改'])


            #处理'24'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '24')
            fun_defect_map['24'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '24', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '24', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '24')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '24', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['24'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '24', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '24', 12, index5, ['整改中', '部分整改'])


            #处理'25'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '25')
            fun_defect_map['25'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '25', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '25', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '25')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '25', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['25'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '25', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '25', 12, index5, ['整改中', '部分整改'])

            #处理'26'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '26')
            fun_defect_map['26'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '26', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '26', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '26')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '26', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['26'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '26', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '26', 12, index5, ['整改中', '部分整改'])

            #处理'30（用扩）
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '30（用扩）')
            fun_defect_map['30（用扩）'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '30（用扩）', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '30（用扩）', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '30（用扩）')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '30（用扩）', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['30（用扩）'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '30（用扩）', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '30（用扩）', 12, index5, ['整改中', '部分整改'])

            #处理'33（银扩）
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '33（银扩）')
            fun_defect_map['33（银扩）'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '33（银扩）', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '33（银扩）', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '33（银扩）')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '33（银扩）', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['33（银扩）'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '33（银扩）', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '33（银扩）', 12, index5, ['整改中', '部分整改'])

            #处理'34（银扩）
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '34（银扩）')
            fun_defect_map['34（银扩）'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '34（银扩）', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '34（银扩）', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '34（银扩）')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '34（银扩）', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['34（银扩）'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '34（银扩）', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '34（银扩）', 12, index5, ['整改中', '部分整改'])

            #处理'35'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '35')
            fun_defect_map['35'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '35', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '35', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '35')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '35', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['35'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '35', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '35', 12, index5, ['整改中', '部分整改'])

            #处理'36'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '36')
            fun_defect_map['36'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '36', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '36', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '36')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '36', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['36'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '36', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '36', 12, index5, ['整改中', '部分整改'])

            #处理'37'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '37')
            fun_defect_map['37'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '37', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '37', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '37')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '37', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['37'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '37', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '37', 12, index5, ['整改中', '部分整改'])

            #处理'38'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '38')
            fun_defect_map['38'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '38', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '38', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '38')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '38', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['38'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '38', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '38', 12, index5, ['整改中', '部分整改'])

            #处理'51'
            index1 += 10
            no_num = func_1(fun_defect, fun_defect_map, index1, '51')
            fun_defect_map['51'][1] += no_num
            index2 += 10
            func_2(fun_defect, fun_defect_map, index2, '51', 2, index1, ['不符合','部分符合'])
            index3 += 10
            func_2(fun_defect, fun_defect_map, index3, '51', 3, index1, ['不符合','部分符合'])
            index4 += 10
            func_3(fun_defect, fun_defect_map, index4, '51')
            index5 += 10
            d_num = func_4(fun_defect, fun_defect_map, index5, '51', 7)
            index8 += 10
            d_num = func_5(fun_defect, index8)
            fun_defect_map['51'][11] += d_num
            index6 += 10
            func_4(fun_defect, fun_defect_map, index6, '51', 8)
            index7 += 10
            func_2(fun_defect, fun_defect_map, index7, '51', 12, index5, ['整改中', '部分整改'])


        fun_defect_ids = []
        for key in fun_defect_map.keys():
            # 处理功能缺陷是否符合字段(合并)
            if key == 'name':
                continue
            array_accord = fun_defect_map[key][0]
            len_accord = array_accord.__len__()
            fun_defect_ids.append(key)
            if len_accord == 0:
                fun_defect_map[key][4] = '不适用'
            elif len_accord == 1:
                fun_defect_map[key][4] = array_accord[0]
            else:
                if '不符合' in array_accord or '部分符合' in array_accord:
                    fun_defect_map[key][4] = '不符合'
                else:
                    fun_defect_map[key][4] = '符合'

            # 处理同意
            array_agree = fun_defect_map[key][5]
            len_agree = array_agree.__len__()
            if len_agree == 0:
                fun_defect_map[key][6] = ''
            elif len_agree == 1:
                fun_defect_map[key][6] = array_agree[0]
            else:
                if '否' in array_agree:
                    fun_defect_map[key][6] = '否'
                else:
                    fun_defect_map[key][6] = '是'

            '''
            ①无须整改
            ②第三方整改
            ③整改中
            ④部分整改
            ⑤已全部整改
            ⑥无法整改
            
            规则：
             --1、所有功能都是唯一的，序号①或②或③或④或⑤或⑥，直接将序号对应值汇总到第一个sheet页；
              --  2、所有功能名称封堵结果是组合情况：优先级别从高到低：④③
                特殊说明：
              --  1>有④，则汇总封堵结果为④
              --  2>若无④，有③，除开④其他序号都有，汇总封堵结果为③;
              --  3>若无④且无③，其他序号都有【①，②，⑤，⑥】，封堵结果为⑤；
              --  4>若无④且无③无②，其他序号都有【①，⑤，⑥】，封堵结果为⑤；
              --   5>若无④且无③且无⑥，其他序号都有【①，②，⑤】，封堵结果为⑤；
              --  6>若无④且无③且无⑤，其他序号【①，②，⑥】，封堵结果为⑥；
                
              --  7>若无④且无③且无⑤无②，其他序号【①，⑥】，封堵结果为⑥；
              --  8>若无④且无③且无⑤无⑥,其他序号【①，②】,封堵结果为②。
              --  9>只有⑤，⑥，封堵汇总结果为⑤。
              --  10>只有①⑤，⑥，封堵汇总结果为⑤。
            '''

            # 处理封堵结果
            array_block = fun_defect_map[key][7]
            len_block = array_block.__len__()
            if len_block == 0:
                fun_defect_map[key][9] = ''
            elif len_block == 1:
                fun_defect_map[key][9] = array_block[0]
            else:
                if '部分整改' in array_block:
                    fun_defect_map[key][9] = '部分整改'
                elif '部分整改' not in array_block and '整改中' in array_block:
                    fun_defect_map[key][9] = '整改中'
                else:
                    valid_set1 = {'无须整改', '第三方整改', '已全部整改', '无法整改'}
                    if '已全部整改' in array_block and all_values_in_set(array_block, valid_set1):
                        fun_defect_map[key][9] = '已全部整改'
                    else:
                        valid_set1 = {'无须整改','无法整改'}
                        valid_set2 = {'无须整改','第三方整改'}
                        valid_set3 = {'无须整改','第三方整改','无法整改'}
                        if '无法整改' in array_block and all_values_in_set(array_block, valid_set3):
                            fun_defect_map[key][9] = '无法整改'
                        else:
                            #all_values_in_set(array_block, valid_set1)
                            #  7>若无④且无③且无⑤无②，其他序号【①，⑥】，封堵结果为⑥；
                            fun_defect_map[key][9] = '第三方整改'


            #处理审核结果
            array_audit = fun_defect_map[key][8]
            len_audit = array_audit.__len__()
            if len_audit == 0:
                fun_defect_map[key][10] = ''
            elif len_audit == 1:
                fun_defect_map[key][10] = array_audit[0]
            else:
                if '不通过' in array_audit:
                    fun_defect_map[key][10] = '不通过'
                else:
                    fun_defect_map[key][10] = '通过'




        for _row in range(maxRow, 0, -1):
            for _col in range(maxCol, 0, -1):
                if _col < _insertColumn:
                    continue

                # print('行：' + str(_row) + '     列：' + str(_col))
                source_cell = sheet.cell(row=_row, column=_col)
                target_cell = sheet.cell(row=_row, column=_col + 1)
                target_cell.value = source_cell.value
                # target_cell._style = copy.copy(source_cell._style)
                # print(1111)
                target_cell.font = copy.copy(source_cell.font)  # 复制字体
                # print(2222)
                target_cell.border = copy.copy(source_cell.border)  # 复制边框
                # print(3333)
                target_cell.fill = copy.copy(source_cell.fill)  # 复制填充
                # print(4444)
                target_cell.number_format = copy.copy(source_cell.number_format)  # 复制数字格式
                # print(5555)
                target_cell.protection = copy.copy(source_cell.protection)  # 复制保护
                # print(6666)
                target_cell.alignment = copy.copy(source_cell.alignment)  # 复制对齐方式
        # 不使用此方法 可能会造成错乱
        # sheet.insert_cols(2, amount=1)
        # 先还原非插入列单元格的合并
        for coord in mergeCellArray:
            min_col, min_row, max_col, max_row = range_boundaries(coord)
            if min_col >= 2:
                min_col = min_col + 1
            if max_col >= 2:
                max_col = max_col + 1
            coordinate = get_column_letter(min_col) + str(min_row) + ":" + get_column_letter(max_col) + str(max_row)
            sheet.merge_cells(coordinate)
            # sheet.unmerge_cells(coordinate)

        # # 先还原非插入列单元格的合并(功能缺陷)
        # for coord in mergeCellArray2:
        #     min_col, min_row, max_col, max_row = range_boundaries(coord)
        #     if min_col >= 2:
        #         min_col = min_col + 1
        #     if max_col >= 2:
        #         max_col = max_col + 1
        #     coordinate = get_column_letter(min_col) + str(min_row) + ":" + get_column_letter(max_col) + str(max_row)
        #     sheet2.merge_cells(coordinate)

        # 还原列宽
        for index, value in enumerate(colWidths):
            sheet.column_dimensions[get_column_letter(index + 4)].width = value

        # # 还原列宽(功能缺陷)
        # for index, value in enumerate(colWidths2):
        #     sheet2.column_dimensions[get_column_letter(index + 4)].width = value
        # 合并多行
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 添加项目名称")
        sheet["B1"].value = "项目名称"
        sheet["B4"].value = projectName
        # 添加边框
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        sheet['B4'].border = border
        sheet['B' + str(maxRow)].border = border
        sheet.merge_cells("B1:B2")
        sheet.merge_cells("B4:B" + str(maxRow))
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet['B1'].alignment = alignment
        sheet['B4'].alignment = alignment
        #  最后因为插入单元格不会把图片也同样位移 因此需要通过代码来移动图片
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 图片位移修正")
        imageArray = []
        for img in sheet._images:
            if img.anchor._from.col >= 2:
                imageArray.append(img)
        for img in imageArray:
            oldFrom = img.anchor._from
            fromImgRow = oldFrom.row
            fromImgCol = oldFrom.col + 1
            fromColOff = oldFrom.colOff
            fromRowOff = oldFrom.rowOff
            _anchor = copy.copy(img.anchor)
            _anchor._from = AnchorMarker(fromImgCol, fromColOff, fromImgRow, fromRowOff)
            if isinstance(img.anchor, TwoCellAnchor):
                oldTo = img.anchor.to
                toImgRow = oldTo.row
                toImgCol = oldTo.col + 1
                toColOff = oldTo.colOff
                toRowOff = oldTo.rowOff
                _anchor.to = AnchorMarker(toImgCol, toColOff, toImgRow, toRowOff)
            else:
                oldExt = img.anchor.ext
                _anchor.ext = XDRPositiveSize2D(oldExt.cx, oldExt.cy)
            # 先删除原来的图片
            del sheet._images[sheet._images.index(img)]
            # 不再使用图片
            # sheet.add_image(img, anchor=_anchor)

        # 遍历所有下拉框，更新列号，然后添加到新位置
        # move_data_validation_to_next_column(sheet)

        #覆盖功能缺陷合并后的值
        for _row in range(3, maxRow, 1):
            row_datas = sheet[_row]
            seq = str(row_datas[0].value)
            if seq in fun_defect_ids:
                # sheet.cell(row=_row, column=8).value = \
                #     ('' if sheet.cell(row=_row, column=8).value is None else sheet.cell(row=_row, column=8).value) + fun_defect_map['name']
                accord = fun_defect_map[seq][4]
                sheet.cell(row=_row, column=9).value = accord
                if not (accord.__len__() == 1 and accord[0] == '符合'):
                    sheet.cell(row=_row, column=10).value = fun_defect_map[seq][1]
                    sheet.cell(row=_row, column=11).value = fun_defect_map[seq][2]
                    sheet.cell(row=_row, column=14).value = fun_defect_map[seq][3]
                sheet.cell(row=_row, column=15).value = fun_defect_map[seq][6]
                sheet.cell(row=_row, column=16).value = fun_defect_map[seq][6]
                sheet.cell(row=_row, column=20).value = fun_defect_map[seq][9]
                sheet.cell(row=_row, column=21).value = fun_defect_map[seq][12]
                sheet.cell(row=_row, column=22).value = fun_defect_map[seq][11]
                sheet.cell(row=_row, column=30).value = fun_defect_map[seq][10]

        wb.save(save_path)
        wb.close()
        # except Exception as e:
        #     wb.save(save_error_path)
        #     wb.close()
        #     print(e)
        #     print(filename + "；处理失败")
        # else:
        #     print(filename + "；处理完成")

    #part 3
    print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 开始合并Excel的系统级隐患清单")
    filenames = os.listdir(excelDirWorkPath)
    # 遍历文件夹下的所有已经预处理好的文件
    mergeWorkBook = None
    mergeSheet = None
    dealFileCount = 0
    startInsertRow = 0
    everySheetStartRow = 4  # 每一个sheet都是从第4行开始复制
    for filename in filenames:
        file_path = os.path.join(excelDirWorkPath, filename)
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 开始合并 " + filename)
        # 读取excel中的第一个sheet
        wb = load_workbook(file_path)
        # 以第一个文件为基础 从第二个文件开始进行合并
        if dealFileCount == 0:
            mergeWorkBook = wb
            mergeSheet = wb[wb.sheetnames[0]]
            startInsertRow = mergeSheet.max_row + 1
            dealFileCount = dealFileCount + 1
            continue
        dealFileCount = dealFileCount + 1
        sheet = wb[wb.sheetnames[0]]
        maxRow = sheet.max_row
        maxColumn = sheet.max_column
        mergeSheet.insert_rows(idx=startInsertRow, amount=maxRow - everySheetStartRow + 1)
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 开始逐一插入合并Sheet中的单元格 ")
        for row in range(everySheetStartRow, maxRow + 1):
            for col in range(1, maxColumn + 1):
                source_cell = sheet.cell(row=row, column=col)
                target_cell = mergeSheet.cell(row=startInsertRow + row - everySheetStartRow, column=col)
                target_cell.value = source_cell.value
                # target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)  # 复制字体
                target_cell.border = copy.copy(source_cell.border)  # 复制边框
                target_cell.fill = copy.copy(source_cell.fill)  # 复制填充
                target_cell.number_format = copy.copy(source_cell.number_format)  # 复制数字格式
                target_cell.protection = copy.copy(source_cell.protection)  # 复制保护
                target_cell.alignment = copy.copy(source_cell.alignment)  # 复制对齐方式
            # 复制行高
            mergeSheet.row_dimensions[startInsertRow + row - everySheetStartRow].height = sheet.row_dimensions[row].height
        # 复制合并单元格
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 处理合并单元格")
        for range_ in sheet.merged_cells.ranges:
            if range_.min_row >= everySheetStartRow:
                min_row = range_.min_row
                max_row = range_.max_row
                min_col = range_.min_col
                max_col = range_.max_col
                # print("minRow:"+str(startInsertRow+min_row-everySheetStartRow)+" maxRow:"+str(startInsertRow+max_row-everySheetStartRow)+" minCol:"+str(min_col)+" maxCol:"+str(max_col))
                # 基于行和列坐标合并单元格
                mergeSheet.merge_cells(start_row=startInsertRow + min_row - everySheetStartRow, start_column=min_col,
                                       end_row=startInsertRow + max_row - everySheetStartRow, end_column=max_col)
        # 获取源单元格中的所有图片
        print(time.strftime('%Y-%m-%d %H:%M:%S') + ": 迁移图片 ")
        for img in sheet._images:
            if img.anchor._from.row >= everySheetStartRow:
                oldFrom = img.anchor._from
                fromImgRow = oldFrom.row + startInsertRow - everySheetStartRow
                fromImgCol = oldFrom.col
                fromColOff = oldFrom.colOff
                fromRowOff = oldFrom.rowOff
                _anchor = copy.copy(img.anchor)
                _anchor._from = AnchorMarker(fromImgCol, fromColOff, fromImgRow, fromRowOff)
                if isinstance(img.anchor, TwoCellAnchor):
                    oldTo = img.anchor.to
                    toImgRow = oldTo.row + startInsertRow - everySheetStartRow
                    toImgCol = oldTo.col
                    toColOff = oldTo.colOff
                    toRowOff = oldTo.rowOff
                    _anchor.to = AnchorMarker(toImgCol, toColOff, toImgRow, toRowOff)
                else:
                    oldExt = img.anchor.ext
                    _anchor.ext = XDRPositiveSize2D(oldExt.cx, oldExt.cy)
                mergeSheet.add_image(img, anchor=_anchor)
        startInsertRow = startInsertRow + maxRow - everySheetStartRow + 1
    mergeWorkBook.save(os.path.join(excelDirWorkPath, "result.xlsx"))

