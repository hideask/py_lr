from datetime import datetime

import pandas as pd
from openpyxl import Workbook
import os
from openpyxl.styles import Font
import requests
import json
import threading
import logging
import time

import logging
import random

import psycopg2
from db_connection import get_db_connection, close_db_connection

# 创建 logger 对象
logger = logging.getLogger()
logger.setLevel(logging.INFO)  # 设置日志级别

# 创建一个用于写入文件的 handler
file_handler = logging.FileHandler('app.log')
file_handler.setLevel(logging.INFO)
file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(file_formatter)

# 创建一个用于输出到控制台的 handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)

# 将两个 handler 添加到 logger 中
logger.addHandler(file_handler)
logger.addHandler(console_handler)


counter = 0


# 结果表中创建新的列
# df = pd.read_excel('C:/Users/liujie/Desktop/resume_data_json_0_5000.xlsx')
# df = pd.read_excel('D:/数字政务/人岗匹配/一句话简历数据/原始数据/resume_data_json_5000_15000.xlsx')
# print(f"读取数据成功，共有{df.shape[0]}条数据")
# column_names = df.columns.tolist()
# column_names.extend(['序号','introduce', 'threadName', '耗时(秒)'])
#
# new_wb = Workbook()
# ws = new_wb.active
# ws.append(column_names)
# for cell in ws[1]:  # 工作表的第一行
#     cell.font = Font(bold=True)  # 设置字体为加粗



# 读取Excel文件
def read_excel():
    global df
    # 每100行进行切片
    slice_size = 482
    slices = [df.iloc[i:i + slice_size] for i in range(0, df.shape[0], slice_size)]
    print(f"处理{len(slices)}个块")
    return slices

check_file_exists = True
# 处理每个块的函数
def process_chunk(chunk, threadName, lock, output_dir):
    # global counter

    thread_output_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}.xlsx')

    column_names = chunk.columns.tolist()
    column_names.extend(['序号','introduce', 'threadName', 'startTime', 'endTime', 'startTime_format', 'endTime_format', '耗时(秒)'])

    # 如果文件不存在，先创建文件并写入表头
    if not os.path.exists(thread_output_file):
        df_header = pd.DataFrame(columns=column_names)
        df_header.to_excel(thread_output_file, index=False)
    elif check_file_exists:
        # 读取已有数据，用于判断 id 是否已存在
        existing_df = pd.read_excel(thread_output_file)
        existing_ids = set(existing_df['id'])  # 假设 'id' 是唯一标识字段

    # results = []
    counter = 1
    for index, row in chunk.iterrows():
        try:
            # 新增：判断 id 是否已存在，若存在则跳过
            if check_file_exists & os.path.exists(thread_output_file) and row['id'] in existing_ids:
                logging.info(f"ID {row['id']} 已存在，跳过处理")
                continue

            user_id = threadName + str(random.randint(1_000_000, 9_999_999))
            AppConversationID = getApplicationId(user_id)
            # AppConversationID = getApplicationId(threadName)
            result = api_call(row, AppConversationID, lock, user_id, counter)


            current_row = row.to_list()
            current_row.extend(result)
            df_row = pd.DataFrame([current_row], columns=column_names)
            with pd.ExcelWriter(thread_output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if 'Sheet1' in writer.book.sheetnames:
                    startrow = writer.sheets['Sheet1'].max_row
                else:
                    startrow = 0
                df_row.to_excel(writer, index=False, header=False, startrow=startrow)

            counter += 1
            if counter % 10 == 0 and counter > 0:
                current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
                logging.info(f"[{current_time}]{threading.current_thread().name} 已处理条数：{counter}")

        except Exception as e:
            logging.error(f"处理数据出错: {str(e)}")
            # 添加错误信息并追加保存
            error_row = row.to_list()
            error_row.extend([counter, f"处理出错: {str(e)}", threading.current_thread().name, 0])
            df_error = pd.DataFrame([error_row], columns=column_names)
            with pd.ExcelWriter(thread_output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if 'Sheet1' in writer.book.sheetnames:
                    startrow = writer.sheets['Sheet1'].max_row
                else:
                    startrow = 0
                df_error.to_excel(writer, index=False, header=False, startrow=startrow)

        # 每处理10条数据保存一次
        # if (index + 1) % 10 == 0:
        #     temp_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}_temp.xlsx')
        #     try:
        #         thread_wb.save(temp_file)
        #         final_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}.xlsx')
        #         if os.path.exists(final_file):
        #             os.remove(final_file)
        #         os.rename(temp_file, final_file)
        #     except Exception as e:
        #         logging.error(f"保存文件时出错: {str(e)}")
        #         if os.path.exists(temp_file):
        #             os.remove(temp_file)
        # 为每个线程保存独立的文件
    # thread_output_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}.xlsx')
    # thread_wb.save(thread_output_file)

    # return results

def format_time(timestamp):
    dt = datetime.fromtimestamp(timestamp)
    millis = int((timestamp - int(timestamp)) * 1000)
    return dt.strftime(f"%Y-%m-%d %H:%M:%S.{millis:03d}")

# 模拟API调用的函数
def api_call(row, AppConversationID, lock, user_id, counter):
    # global counter
    start_time = time.time()  # 开始计时

    url = 'http://10.163.21.201:32300/api/proxy/api/v1/chat_query'
    data = {
        'Apikey': 'd13tc102gkoadkll1qgg',
        "Query": row['info'],
        "AppConversationID": AppConversationID,
        "ResponseMode": "blocking",
        "UserID": user_id
    }

    headers = {
        'Content-Type': 'application/json',
        'Apikey': 'd13tc102gkoadkll1qgg',
    }

    response = requests.post(url, json=data, headers=headers)

    introduce = '无'

    if response.status_code == 200:
        start = response.text.index('{')
        result = json.loads(response.text[start:-1]).get("answer")
        result = result.encode('ISO_8859_1').decode('utf-8')
        try:
            if result != '':
                introduce = result
            else:
                error = json.loads(result[result.index('{'):result.index('}') + 1]).get('error')
                logging.error(f'error:{error}')
        except Exception as e:
            logging.error(f'数据解析失败:{e}')
    else:
        logging.error(f'Failed to post data:{response.text}')

    end_time = time.time()  # 结束计时
    elapsed_time = round(end_time - start_time, 2)  # 精确到小数点后两位

    # lock.acquire()
    # try:
    #     counter += 1
        # logging.info(f"处理数据, user_id: {user_id}")
        # logging.info(f'counter:{counter}')
        # logging.info(f"id:{row['id']}")
        # logging.info(f'耗时:{elapsed_time}秒')
        # logging.info('------------------------------')

    # finally:
    #     lock.release()
    return [counter,introduce, threading.current_thread().name,start_time, end_time, format_time(start_time), format_time(end_time), elapsed_time]


def getApplicationId(user_id):
    url = 'http://10.163.21.201:32300/api/proxy/api/v1/create_conversation'
    data1 = {
        'Apikey': 'd13tc102gkoadkll1qgg',
        "Inputs": {
            "var": "variable"
        },
        "UserID": user_id
    }

    headers = {
        'Content-Type': 'application/json',
        'Apikey': 'd13tc102gkoadkll1qgg'
    }

    response1 = requests.post(url, json=data1, headers=headers)

    if response1.status_code == 200:
        AppConversationID = response1.json().get("Conversation").get('AppConversationID')
    else:
        AppConversationID = ''
        print('Failed to post data:', response1.status_code)
    return AppConversationID


# 主函数
def main(output_dir):
    global new_wb

    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 读取Excel文件
    excel_chunks = read_excel()

    threads = []
    lock = threading.Lock()
    for index, chunk in enumerate(excel_chunks):
        threadName = f'AAThread-{index}'

        thread = threading.Thread(target=process_chunk, args=(chunk, threadName, lock,  output_dir), name=threadName)
        threads.append(thread)
        thread.start()
    # 等待所有线程完成
    for thread in threads:
        thread.join()

    # 将结果写入新的Excel文件
    # output_file = os.path.join(output_dir, '智能体分发结果.xlsx')
    # new_wb.save(output_file)
    print("执行完毕")


def fetch_unprocessed_data_from_db(batch_size=100):
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # 查询未处理的数据并标记为正在处理
        cursor.execute("""
            SELECT id, resume_processed_info as info FROM zhilian_resume 
            WHERE resume_description_detail IS NULL 
            and process_type is null
            and length(resume_processed_info) > 5000 and length(resume_processed_info) <= 15000 and id != '25048' 
            ORDER BY id LIMIT %s FOR UPDATE SKIP LOCKED
        """, (batch_size,))

        rows = cursor.fetchall()

        if rows:
            ids = [row[0] for row in rows]
            # 标记这些数据为正在处理
            cursor.execute("""
                UPDATE zhilian_resume SET process_type = %s 
                WHERE id = ANY(%s)
            """, (threading.current_thread().name,ids,))
            connection.commit()
        return rows

    except Exception as e:
        logging.error(f"Error fetching data from DB: {e}")
        connection.rollback()
    finally:
        close_db_connection(cursor, connection)
    return []

def update_result_in_db(id, result):
    connection = get_db_connection()
    cursor = connection.cursor()
    # return [counter,introduce, threading.current_thread().name,start_time, end_time, format_time(start_time), format_time(end_time), elapsed_time]

    try:
        introduce, start_time, end_time, elapsed_time = result[1], result[5], result[6], result[7]
        cursor.execute("""
            UPDATE zhilian_resume SET 
                resume_description_detail = %s,
                process_start_time = %s,
                process_end_time = %s,
                time_consume = %s,
                process_type = '1'
            WHERE id = %s
        """, (introduce, start_time, end_time, elapsed_time, id))
        connection.commit()
        logging.info(f"线程:{threading.current_thread().name}已更新 id：{id}...耗时：{elapsed_time}")
    except Exception as e:
        logging.error(f"Error updating database for ID {id}: {e}")
        connection.rollback()
    finally:
        close_db_connection(cursor, connection)

def worker_thread(lock, batch_size=100):
    while True:
        with lock:
            rows = fetch_unprocessed_data_from_db(batch_size)
        if not rows:
            break  # 没有更多数据了，退出循环
        logging.info(f"线程:{threading.current_thread().name}正在处理 {len(rows)} 条数据...")
        for row in rows:
            id = row[0]  # 第一个字段是 id
            info = row[1]  # 第二个字段是 resume_processed_info
            call_data = {'id': id, 'info': info}
            try:
                user_id = threading.current_thread().name + str(random.randint(1_000_000, 9_999_999))
                AppConversationID = getApplicationId(user_id)
                # AppConversationID = getApplicationId(threadName)
                result = api_call(call_data, AppConversationID, lock, user_id, counter)
                update_result_in_db(id, result)  # 将结果写回数据库
            except Exception as e:
                logging.error(f"Error processing ID {id}: {e}")
                # 可以在这里选择重新放回未处理状态或跳过

def start_processing_with_threads(num_threads=20, batch_size=10):
    lock = threading.Lock()
    threads = []

    for i in range(num_threads):
        thread = threading.Thread(target=worker_thread, args=(lock, batch_size), name=f"Worker-{i+1}")
        threads.append(thread)
        thread.start()
    for thread in threads:
        thread.join()

    print("All data has been processed.")


from concurrent.futures import ThreadPoolExecutor, as_completed


def start_with_restarting_pool(num_threads=20, max_retries=5, batch_size=10):
    lock = threading.Lock()

    def run_task(retry_count=0):
        if retry_count > max_retries:
            logging.warning("超过最大重试次数，放弃任务")
            return
        try:
            worker_thread(lock, batch_size)
        except Exception as e:
            logging.error(f"任务失败，第 {retry_count} 次重试: {e}")
            run_task(retry_count + 1)

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(run_task, 0) for _ in range(num_threads)]

        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logging.error(f"任务执行失败: {e}")

def get_excel_files(folder_path):
    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".xlsx") or file.endswith(".xls"):
                excel_files.append(os.path.join(root, file))
    return excel_files

# 处理单个Excel文件并更新数据
def process_excel_file(file_path):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        df = pd.read_excel(file_path)
        # 假设有一个'id'字段用于匹配和更新
        for index, row in df.iterrows():
            # 这里可以根据你的需求进行数据库更新或其他操作
            cursor.execute("update zhilian_resume set resume_description_detail = %s, process_start_time = %s,"
                           "process_end_time = %s,time_consume = %s, process_type = %s where id = %s",
                           (row['introduce'], time_to_timestamp(row['startTime_format'],  date_str='2025-06-10'), time_to_timestamp(row['endTime_format'], date_str='2025-06-10'), row['耗时(秒)'], '1', row['id']))
            connection.commit()
            print(f"更新数据成功, id: {row['id']}")
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
    finally:
        close_db_connection(cursor, connection)


def time_to_timestamp(time_str, date_str=None):
    if date_str is None:
        date_part = datetime.today().date()
    else:
        date_part = datetime.strptime(date_str, "%Y-%m-%d").date()

    # 分离时间和毫秒
    time_parts = time_str.split('.')
    main_time = time_parts[0]
    millis = int(time_parts[1]) if len(time_parts) > 1 else 0

    # 合并日期和时间
    datetime_str = f"{date_part} {main_time}"
    datetime_obj = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")

    # 添加毫秒
    datetime_obj_with_ms = datetime_obj.replace(microsecond=millis * 1000)

    # 返回 PostgreSQL 支持的时间字符串格式
    return datetime_obj_with_ms.strftime("%Y-%m-%d %H:%M:%S.%f")

# 线程任务：每个线程处理两个文件
def thread_task(files):
    for file in files:
        process_excel_file(file)

def update_resume_main():
    file_path = 'D:/数字政务/人岗匹配/一句话简历数据/处理数据/线程文件/0-5000'
    files = get_excel_files(file_path)
    file_groups = [files[i:i + 2] for i in range(0, len(files), 2)]
    threads = []
    for group in file_groups:
        thread = threading.Thread(target=thread_task, args=(group,))
        threads.append(thread)
        thread.start()

    for  thread in threads:
        thread.join()

def update_resume_main_sc():
    connection = get_db_connection()
    cursor = connection.cursor()
    fiile_path = 'D:/数字政务/人岗匹配/一句话简历数据/处理数据/人岗匹配_sc_0_13000_20250609092438.xlsx'
    df = pd.read_excel(fiile_path)
    for index, row in df.iterrows():
        input = row['input']
        modelAnswer = row['modelAnswer']
        start_marker = "```json\n"
        end_marker = "```"
        json_start = input.find(start_marker) + len(start_marker)
        json_end = input.find(end_marker, json_start)
        if json_start == -1 or json_end == -1:
            print(f"第{index + 1}条数据没有找到json数据")
            continue

        json_data = json.loads(input[json_start:json_end].strip())
        id = json_data['id']
        cursor.execute("update sc_pub_recruitmentnet_resume set resume_description_detail = %s where id = %s",
                       (modelAnswer, id))
        connection.commit()
        print(f"更新数据成功, id: {id}")

# 调用主函数
if __name__ == "__main__":
    # output_dir = 'D:/数字政务/人岗匹配/一句话简历数据/处理数据/线程文件/end4000'  # 输出目录
    # main(output_dir)
    # update_resume_main()
    update_resume_main_sc()

    # num_threads = 10
    # batch_size = 10
    # # start_processing_with_threads(num_threads=num_threads, batch_size=batch_size)
    # max_retries = 5
    # start_with_restarting_pool(num_threads=num_threads,max_retries=5, batch_size=batch_size)