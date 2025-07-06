import re
import psycopg2
from datetime import datetime
from db_connection import get_db_connection, close_db_connection

import pandas as pd
from openpyxl import Workbook
import os
from openpyxl.styles import Font
import requests
import json
import threading
import logging
import time
import random

import logging

logging.basicConfig(level=logging.INFO)

retain_fields = {
    "companyName": "",
    "cityDistrict": "",
    "education": "",
    "jobSummary": "",
    "name": "",
    "recruitNumber": True,
    "salaryReal": "",
    "welfareTagList": [],
    "workType": "",
    "workingExp": ""
    # "companySize": "",
    # "displayPhoneNumber": False,
    # "industryName": "",
    # "cardCustomJson": "",
    # "needMajor": [],
    # "staffCard": {
    #     "staffName": ""
    # },
    # "salary60": "",
    # "salaryCount": "",
    # "subJobTypeLevelName": "",
    # "workCity": "",
    # "streetName": "",
}


import html

def clean_html(text):
    if not text:
        return ""

    # 第一步：自动解码 HTML 实体（如 &lt; → <, &#xa; → \n, &#xff01; → ！）
    text = html.unescape(text)

    # 第二步：去除所有 HTML 标签
    text = re.sub(r"<[^>]+>", "", text)

    text = re.sub(r"&\s*lt\s*;?", "<", text)
    text = re.sub(r"&\s*gt\s*;?", ">", text)
    text = re.sub(r"&\s*#xa\s*;?", "\n", text)
    text = re.sub(r"&\s*#xd\s*;?", "\r", text)
    text = re.sub(r"&\s*#x9\s*;?", " ", text)

    # 第三步：替换特殊空白字符（零宽空格、软换行等）
    text = re.sub(r"[\xa0\u200b\u200c\u200d\u200e\u200f]", " ", text)

    # 第四步：合并连续空白为单个空格，并去除首尾空格
    text = re.sub(r"\s+", " ", text).strip()

    return text

def clean_job_data():
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT id, job_description FROM sc_pub_recruitmentnet_job where job_description like '%<%' or job_description like '%&%' ")

    rows = cursor.fetchall()

    for row in rows:
        job_id = row[0]
        dirty_description = row[1]
        cleaned_description = clean_html(dirty_description)

        cursor.execute(
            "UPDATE sc_pub_recruitmentnet_job SET job_description = %s WHERE id = %s",
            (cleaned_description, job_id)
        )
        connection.commit()
        print(f"Cleaned job description for job ID: {job_id}")
    close_db_connection(cursor, connection)


def process_jobs():
    """处理岗位主函数"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        zhilian_job = process_job_batch(cursor, connection)

        i = 0
        for id, job_info, category_name, job_name in zhilian_job:
            try:
                job_data = json.loads(job_info)
                filtered_data = filter_job_data(job_data, required_fields=retain_fields)
                
                # 重组JSON字段：提取cardCustomJson中的字段到最外层
                if 'cardCustomJson' in job_data and job_data['cardCustomJson']:
                    try:
                        card_custom_data = json.loads(job_data['cardCustomJson'])
                        # 提取salary60到最外层
                        # if 'salary60' in card_custom_data:
                        #     filtered_data['salary60'] = card_custom_data['salary60']
                        #     del card_custom_data['salary60']

                        # 提取address到最外层
                        if 'address' in card_custom_data:
                            filtered_data['address'] = card_custom_data['address']
                            del card_custom_data['address']
                        # 更新cardCustomJson，移除已提取的字段
                        # filtered_data['cardCustomJson'] = json.dumps(card_custom_data, ensure_ascii=False)
                    except (json.JSONDecodeError, TypeError):
                        # 如果cardCustomJson不是有效的JSON，保持原样
                        pass
                
                # 提取staffCard中的staffName到最外层
                if 'staffCard' in filtered_data and isinstance(filtered_data['staffCard'], dict):
                    if 'staffName' in filtered_data['staffCard']:
                        filtered_data['staffName'] = filtered_data['staffCard']['staffName']
                    # 移除staffCard对象
                    del filtered_data['staffCard']
                
                # filtered_data['category_name'] = category_name
                # filtered_data['job_name'] = job_name
                filtered_data['id'] = id

                # 处理salaryReal字段格式
                if 'salaryReal' in filtered_data:
                    filtered_data['salaryReal'] = process_salary_range(filtered_data['salaryReal'])
                processed_info = json.dumps(filtered_data, ensure_ascii=False)

                update_job_info(cursor, connection, id, processed_info)
                i += 1
                print(f"已更新岗位id {filtered_data['id']}, 成功更新{i}条")
            except Exception as ex:
                connection.rollback()
                print(f"处理岗位id {filtered_data['id']} 时出错: {str(ex)}")
                # print(f"错误数据: {job_info}")
                continue
    finally:
        close_db_connection(cursor, connection)

def process_salary_range(salary_range):
    """处理薪资范围格式，将类似5001-10000转换为5000-10000"""
    if not salary_range or not isinstance(salary_range, str):
        return salary_range
    
    # 匹配薪资范围格式，如：5001-10000, 10001-25000
    pattern = r'(\d+)-(\d+)'
    match = re.match(pattern, salary_range.strip())
    
    if match:
        min_salary = int(match.group(1))
        max_salary = int(match.group(2))
        
        # 将最小薪资调整为千位整数
        # 如果是x001格式，改为x000
        if min_salary % 1000 == 1:
            min_salary = min_salary - 1
        
        return f"{min_salary}-{max_salary}"
    
    return salary_range

def process_job_batch(cursor, connection, batch_size=20000):
    """处理一批岗位数据"""
    cursor.execute("SELECT id, job_info, category_name, job_name FROM zhilian_job where train_type='3' LIMIT %s",
                  (batch_size,))
    zhilian_job = cursor.fetchall()
    print(f"获取成功{len(zhilian_job)}条数据")
    return zhilian_job

def filter_job_data(input_data, required_fields):
    ""
    def filter_dict(data, fields):
        if not isinstance(data, dict):
            return data
        result = {}
        for key, value in data.items():
            if key in fields:
                value = deep_clean(value)
                current_key = key
                if isinstance(value, dict):
                    result[current_key] = filter_dict(value, fields.get(key, {}))
                elif isinstance(value, list):
                    result[current_key] = [filter_dict(item, fields.get(key, {}))
                            if isinstance(item, dict) else item for item in value
                    ]
                else:
                    result[current_key] = value
        return result

    return filter_dict(input_data, required_fields)

def update_job_info(cursor, connection, id, processed_info):
    """更新岗位处理信息"""
    update_sql = """UPDATE zhilian_job SET processed_info = %s WHERE id = %s"""
    cursor.execute(update_sql, (processed_info, id))
    connection.commit()

def deep_clean(data):
    if isinstance(data, str):
        data = re.sub(r'[\u200b-\u200f\u202c-\u202e\ufeff]', '', data)
        # 如果检测到编号列表，则用空格替换换行符，否则删除
        # if re.search(r'\d+\.', data):  # 检测是否有编号
        #     data = data.replace('\n', ' ')
        # else:
        #     data = data.replace('\n', '')
        return data
    elif isinstance(data, dict):
        return {k: deep_clean(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [deep_clean(i) for i in data]
    else:
        return data

def process_sc_jobs():
    """处理岗位主函数"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        cursor.execute("SELECT id,company_name,company_type,industry,company_size,registration_address,job_category,position_name,"
                       "job_description,work_location,required_staff_count,salary_range,benefits,experience_requirement,"
                       "employment_type FROM sc_pub_recruitmentnet_job ")
        zhilian_job = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        i = 0
        for job_info in zhilian_job:
            try:
                row_dict = dict(zip(columns, job_info))
                processed_info = json.dumps(row_dict, ensure_ascii=False)

                update_sql = """UPDATE sc_pub_recruitmentnet_job SET processed_info = %s WHERE id = %s"""
                cursor.execute(update_sql, (processed_info, row_dict['id']))
                connection.commit()
                i += 1
                print(f"已更新岗位id {job_info[0]}, 成功更新{i}条")
            except Exception as ex:
                connection.rollback()
                print(f"处理岗位id {job_info[0]} 时出错: {str(ex)}")
                # print(f"错误数据: {job_info}")
                continue
    finally:
        close_db_connection(cursor, connection)

def process_job_excel():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # cursor.execute(f"select id,processed_info as info  from zhilian_job  where process_type IS null order by id")
        cursor.execute(f"select id,processed_info as info  from sc_pub_recruitmentnet_job  where process_type IS null order by id")
        jobs = cursor.fetchall()
        print(f"获取成功{len(jobs)}条数据")
        # 创建数据列表
        excel_data = []

        # 遍历数据并组装
        for id, info in jobs:
            info = json.loads(info)
            info['id'] = id

            row_data = {
                'id': id,
                'info': json.dumps(info, ensure_ascii=False)
            }
            excel_data.append(row_data)

        df = pd.DataFrame(excel_data)

        output_file = f'job_data_json_{len(jobs)}.xlsx'
        df.to_excel(output_file, index=False)
        print(f"数据已保存到 {output_file}")
    finally:
        close_db_connection(cursor, connection)

check_file_exists = False
def process_chunk(chunk, threadName, lock, output_dir):
    # global counter

    thread_output_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}.xlsx')

    column_names = chunk.columns.tolist()
    column_names.extend(['序号','introduce', 'threadName', 'startTime', 'endTime', 'startTime_format', 'endTime_format', '耗时(秒)'])

    # 如果文件不存在，先创建文件并写入表头
    if not os.path.exists(thread_output_file):
        df_header = pd.DataFrame(columns=column_names)
        df_header.to_excel(thread_output_file, index=False)
    elif check_file_exists:
        # 读取已有数据，用于判断 id 是否已存在
        existing_df = pd.read_excel(thread_output_file)
        existing_ids = set(existing_df['id'])  # 假设 'id' 是唯一标识字段

    # results = []
    counter = 1
    for index, row in chunk.iterrows():
        try:
            # 新增：判断 id 是否已存在，若存在则跳过
            if check_file_exists & os.path.exists(thread_output_file) and row['id'] in existing_ids:
                logging.info(f"ID {row['id']} 已存在，跳过处理")
                continue

            user_id = threadName + str(random.randint(1_000_000, 9_999_999))
            AppConversationID = getApplicationId(user_id)
            # AppConversationID = getApplicationId(threadName)
            result = api_call(row, AppConversationID, lock, user_id, counter)


            current_row = row.to_list()
            current_row.extend(result)
            df_row = pd.DataFrame([current_row], columns=column_names)
            with pd.ExcelWriter(thread_output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if 'Sheet1' in writer.book.sheetnames:
                    startrow = writer.sheets['Sheet1'].max_row
                else:
                    startrow = 0
                df_row.to_excel(writer, index=False, header=False, startrow=startrow)

            counter += 1
            if counter % 10 == 0 and counter > 0:
                current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
                logging.info(f"[{current_time}]{threading.current_thread().name} 已处理条数：{counter}")

        except Exception as e:
            logging.error(f"处理数据出错: {str(e)}")
            # 添加错误信息并追加保存
            error_row = row.to_list()
            error_row.extend([counter, f"处理出错: {str(e)}", threading.current_thread().name, 0])
            df_error = pd.DataFrame([error_row], columns=column_names)
            with pd.ExcelWriter(thread_output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                if 'Sheet1' in writer.book.sheetnames:
                    startrow = writer.sheets['Sheet1'].max_row
                else:
                    startrow = 0
                df_error.to_excel(writer, index=False, header=False, startrow=startrow)

        # 每处理10条数据保存一次
        # if (index + 1) % 10 == 0:
        #     temp_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}_temp.xlsx')
        #     try:
        #         thread_wb.save(temp_file)
        #         final_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}.xlsx')
        #         if os.path.exists(final_file):
        #             os.remove(final_file)
        #         os.rename(temp_file, final_file)
        #     except Exception as e:
        #         logging.error(f"保存文件时出错: {str(e)}")
        #         if os.path.exists(temp_file):
        #             os.remove(temp_file)
        # 为每个线程保存独立的文件
    # thread_output_file = os.path.join(output_dir, f'智能体分发结果_{threading.current_thread().name}.xlsx')
    # thread_wb.save(thread_output_file)

    # return results
def read_excel():
    df = pd.read_excel('D:/数字政务/人岗匹配/一句话简历数据/原始数据/job_data_json_9134.xlsx')
    print(f"读取数据成功，共有{df.shape[0]}条数据")
    # 每100行进行切片
    slice_size = 914
    slices = [df.iloc[i:i + slice_size] for i in range(0, df.shape[0], slice_size)]
    print(f"处理{len(slices)}个块")
    return slices
def format_time(timestamp):
    dt = datetime.fromtimestamp(timestamp)
    millis = int((timestamp - int(timestamp)) * 1000)
    return dt.strftime(f"%Y-%m-%d %H:%M:%S.{millis:03d}")

counter = 0
# 模拟API调用的函数
def api_call(row, AppConversationID, lock, user_id, counter):
    # global counter
    start_time = time.time()  # 开始计时

    url = 'http://10.163.21.201:32300/api/proxy/api/v1/chat_query'
    data = {
        'Apikey': 'd14l347sbfv9olu4aerg',
        "Query": row['info'],
        "AppConversationID": AppConversationID,
        "ResponseMode": "blocking",
        "UserID": user_id
    }

    headers = {
        'Content-Type': 'application/json',
        'Apikey': 'd14l347sbfv9olu4aerg',
    }

    response = requests.post(url, json=data, headers=headers)

    introduce = '无'

    if response.status_code == 200:
        start = response.text.index('{')
        result = json.loads(response.text[start:-1]).get("answer")
        result = result.encode('ISO_8859_1').decode('utf-8')
        try:
            if result != '':
                introduce = result
            else:
                error = json.loads(result[result.index('{'):result.index('}') + 1]).get('error')
                logging.error(f'error:{error}')
        except Exception as e:
            logging.error(f'数据解析失败:{e}')
    else:
        logging.error(f'Failed to post data:{response.text}')

    end_time = time.time()  # 结束计时
    elapsed_time = round(end_time - start_time, 2)  # 精确到小数点后两位

    # lock.acquire()
    # try:
    #     counter += 1
        # logging.info(f"处理数据, user_id: {user_id}")
        # logging.info(f'counter:{counter}')
        # logging.info(f"id:{row['id']}")
        # logging.info(f'耗时:{elapsed_time}秒')
        # logging.info('------------------------------')

    # finally:
    #     lock.release()
    return [counter,introduce, threading.current_thread().name,start_time, end_time, format_time(start_time), format_time(end_time), elapsed_time]


def getApplicationId(user_id):
    url = 'http://10.163.21.201:32300/api/proxy/api/v1/create_conversation'
    data1 = {
        'Apikey': 'd14l347sbfv9olu4aerg',
        "Inputs": {
            "var": "variable"
        },
        "UserID": user_id
    }

    headers = {
        'Content-Type': 'application/json',
        'Apikey': 'd14l347sbfv9olu4aerg'
    }

    response1 = requests.post(url, json=data1, headers=headers)

    if response1.status_code == 200:
        AppConversationID = response1.json().get("Conversation").get('AppConversationID')
    else:
        AppConversationID = ''
        print('Failed to post data:', response1.status_code)
    return AppConversationID

def job_main(output_dir):
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 读取Excel文件
    excel_chunks = read_excel()

    threads = []
    lock = threading.Lock()
    for index, chunk in enumerate(excel_chunks):
        threadName = f'AAThread-{index}'

        thread = threading.Thread(target=process_chunk, args=(chunk, threadName, lock,  output_dir), name=threadName)
        threads.append(thread)
        thread.start()
    # 等待所有线程完成
    for thread in threads:
        thread.join()

    # 将结果写入新的Excel文件
    # output_file = os.path.join(output_dir, '智能体分发结果.xlsx')
    # new_wb.save(output_file)
    print("执行完毕")


def fetch_unprocessed_data_from_db(batch_size=100):
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # 查询未处理的数据并标记为正在处理
        cursor.execute("""
            SELECT id, processed_info as info FROM zhilian_job 
            WHERE job_description_detail IS NULL 
            and process_type is null
            ORDER BY id LIMIT %s FOR UPDATE SKIP LOCKED
        """, (batch_size,))

        rows = cursor.fetchall()

        if rows:
            ids = [row[0] for row in rows]
            # 标记这些数据为正在处理
            cursor.execute("""
                UPDATE zhilian_job SET process_type = %s 
                WHERE id = ANY(%s)
            """, (threading.current_thread().name,ids,))
            connection.commit()
        return rows

    except Exception as e:
        logging.error(f"Error fetching data from DB: {e}")
        connection.rollback()
    finally:
        close_db_connection(cursor, connection)
    return []

def update_result_in_db(id, result):
    connection = get_db_connection()
    cursor = connection.cursor()
    # return [counter,introduce, threading.current_thread().name,start_time, end_time, format_time(start_time), format_time(end_time), elapsed_time]

    try:
        introduce, start_time, end_time, elapsed_time = result[1], result[5], result[6], result[7]
        cursor.execute("""
            UPDATE zhilian_job SET 
                job_description_detail = %s,
                process_start_time = %s,
                process_end_time = %s,
                time_consume = %s,
                process_type = '1'
            WHERE id = %s
        """, (introduce, start_time, end_time, elapsed_time, id))
        connection.commit()
        logging.info(f"线程:{threading.current_thread().name}已更新 id：{id}...耗时：{elapsed_time}")
    except Exception as e:
        logging.error(f"Error updating database for ID {id}: {e}")
        connection.rollback()
    finally:
        close_db_connection(cursor, connection)

def worker_thread(lock, batch_size=100):
    while True:
        with lock:
            rows = fetch_unprocessed_data_from_db(batch_size)
        if not rows:
            break  # 没有更多数据了，退出循环
        logging.info(f"线程:{threading.current_thread().name}正在处理 {len(rows)} 条数据...")
        for row in rows:
            id = row[0]  # 第一个字段是 id
            info = row[1]  # 第二个字段是 processed_info
            call_data = {'id': id, 'info': info}
            try:
                user_id = threading.current_thread().name + str(random.randint(1_000_000, 9_999_999))
                AppConversationID = getApplicationId(user_id)
                # AppConversationID = getApplicationId(threadName)
                result = api_call(call_data, AppConversationID, lock, user_id, counter)
                update_result_in_db(id, result)  # 将结果写回数据库
            except Exception as e:
                logging.error(f"Error processing ID {id}: {e}")
                # 可以在这里选择重新放回未处理状态或跳过

def start_processing_with_threads(num_threads=20, batch_size=10):
    lock = threading.Lock()
    threads = []

    for i in range(num_threads):
        thread = threading.Thread(target=worker_thread, args=(lock, batch_size), name=f"Worker-{i+1}")
        threads.append(thread)
        thread.start()
    for thread in threads:
        thread.join()

    print("All data has been processed.")


from concurrent.futures import ThreadPoolExecutor, as_completed


def start_with_restarting_pool(num_threads=20, max_retries=5, batch_size=10):
    lock = threading.Lock()

    def run_task(retry_count=0):
        if retry_count > max_retries:
            logging.warning("超过最大重试次数，放弃任务")
            return
        try:
            worker_thread(lock, batch_size)
        except Exception as e:
            logging.error(f"任务失败，第 {retry_count} 次重试: {e}")
            run_task(retry_count + 1)

    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(run_task, 0) for _ in range(num_threads)]

        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logging.error(f"任务执行失败: {e}")

def update_job_main():
    file_path = 'D:/数字政务/人岗匹配/一句话简历数据/处理数据/线程文件/job9134'
    files = get_excel_files(file_path)
    file_groups = [files[i:i + 2] for i in range(0, len(files), 2)]
    threads = []
    for group in file_groups:
        thread = threading.Thread(target=thread_task, args=(group,))
        threads.append(thread)
        thread.start()

    for  thread in threads:
        thread.join()

def get_excel_files(folder_path):
    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".xlsx") or file.endswith(".xls"):
                excel_files.append(os.path.join(root, file))
    return excel_files

def thread_task(files):
    for file in files:
        process_excel_file(file)

def process_excel_file(file_path):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        df = pd.read_excel(file_path)
        # 假设有一个'id'字段用于匹配和更新
        for index, row in df.iterrows():
            # 这里可以根据你的需求进行数据库更新或其他操作
            cursor.execute("update zhilian_job set job_description_detail = %s, process_start_time = %s,"
                           "process_end_time = %s,time_consume = %s, process_type = %s where id = %s",
                           (row['introduce'], row['startTime_format'], row['endTime_format'], row['耗时(秒)'], '1', row['id']))
            connection.commit()
            print(f"更新数据成功, id: {row['id']}")
    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
    finally:
        close_db_connection(cursor, connection)

if __name__ == "__main__":
    # clean_job_data()
    process_jobs()
    # process_sc_jobs()
    # process_job_excel()
    # output_dir = 'D:/数字政务/人岗匹配/一句话简历数据/处理数据/线程文件/job9134'  # 输出目录
    # job_main(output_dir)

    # update_job_main()

    # num_threads = 20
    # batch_size = 10
    # max_retries = 5
    # start_with_restarting_pool(num_threads=num_threads,max_retries=5, batch_size=batch_size)